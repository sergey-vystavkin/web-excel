import logging
from datetime import datetime
from os import path
from re import sub

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class Input:

    def __init__(self, config, file_path):
        self.config = config
        self.path = file_path
        self.header = []
        self.header_idx = None
        self.has_non_processed = False
        self.input_sheet = config.get('settings', 'input_sheet')
        self.process_sheet = config.get('settings', 'process_sheet')
        self.no_process_sheet = config.get('settings', 'non_process_sheet')

    def __rename_sheet(self):
        wb = load_workbook(self.path)
        wb.worksheets[0].title = self.input_sheet
        wb.save(self.path)

    def __get_header(self, input_df):
        col_names = dict(
            (k, sub("\\s|\\n", "", self.config.get('input_cols', k).lower())) for k in self.config['input_cols'].keys())
        if len(input_df.columns) < len(col_names):
            raise Exception
        for row_idx, row in input_df.iterrows():
            row_values = [sub("\\s|\\n", "", str(col).lower().strip()) for col in row.values]
            columns_dict = {}

            for key in col_names.keys():
                if col_names[key] not in row_values:
                    break
                for idx, value in enumerate(row_values):
                    if col_names[key] == value:
                        columns_dict[key] = idx
                        break

            if len(columns_dict) == len(col_names):
                self.header = tuple(row.values)
                self.header_idx = row_idx
                return columns_dict  # columns_dict = {config_key: column_idx}
        raise Exception

    def __wrong_file_exception(self, message, reason):
        logger.warning(self.config.get('warning', reason).format(message.subject))
        message.fail_reply(reason=reason)
        raise Exception('*Failed_message')

    def __read_input(self, message):
        try:
            read_wb = pd.ExcelFile(self.path)
        except Exception as exception:
            if exception.args[0] in ['Workbook is encrypted', "Can't find workbook in OLE2 compound document"]:
                self.__wrong_file_exception(message, reason='password_protect')
            else:
                self.__wrong_file_exception(message, reason='cant_open')

        if len(read_wb.sheet_names) > 1:
            self.__wrong_file_exception(message, reason='more_sheets')

        with pd.ExcelFile(self.path) as reader:
            input_df = pd.read_excel(reader, read_wb.sheet_names[0], header=None, keep_default_na=False,
                                     na_filter=False)

        if len(input_df.index.values) == 0:
            self.__wrong_file_exception(message, reason='no_data')
        try:
            self.columns_dict = self.__get_header(input_df)
        except Exception:
            self.__wrong_file_exception(message, reason='wrong_headers')
        return input_df

    def __add_sheets_to_input(self, procces_df, non_process_df, input_df):
        try:  # append sheets
            self.__rename_sheet()
            with pd.ExcelWriter(self.path, mode='a') as writer:
                if len(procces_df.index) > 1:
                    procces_df.to_excel(writer, self.process_sheet, index=False, header=False)
                if len(non_process_df.index) > 1:
                    non_process_df.to_excel(writer, self.no_process_sheet, index=False, header=False)
                    self.has_non_processed = True

        except Exception:
            try:  # rewrite file
                with pd.ExcelWriter(self.path, mode='w') as writer:
                    input_df.to_excel(writer, self.input_sheet, index=False, header=False)
                    if len(procces_df.index) > 1:
                        procces_df.to_excel(writer, self.process_sheet, index=False, header=False)
                    if len(non_process_df.index) > 1:
                        non_process_df.to_excel(writer, self.no_process_sheet, index=False, header=False)
                        self.has_non_processed = True
            except Exception as e:
                logger.error(self.config.get('error', 'write_input').format(path.basename(self.path)))
                raise e

    def parse_input(self, message):
        logger.info(self.config.get('info', 'parse_input'))
        input_df = self.__read_input(message)

        df_without_header = input_df.drop(list(range(self.header_idx + 1)))
        if df_without_header.empty:
            self.__wrong_file_exception(message, reason='no_data')

        process = []
        non_process = []
        for row_idx, input_row in df_without_header.iterrows():
            row = input_row.copy()
            if all(val.strip() == '' for val in row):
                continue  # Skip empty row

            # Delete all symbols besides number and letter  from Contract and Task number Columns
            for col_idx in [self.columns_dict['contract_num'], self.columns_dict['contract_num']]:
                row[col_idx] = sub('\\W', '', row[col_idx])

            row_conv = [sub("\\s|\\n", "", str(val).lower().strip()) for val in row]

            if not pd.isna(row[self.columns_dict['search_name']]) and row[
                self.columns_dict['search_name']].strip() != '':
                process.append(tuple(row.values))
                continue

            for condition in self.config.options('parse_conditions'):
                if eval(condition, {'r_c': row_conv, 'col': self.columns_dict}):
                    search_name = eval(self.config.get('parse_conditions', condition), {'r': row, 'col': self.columns_dict})
                    row[self.columns_dict['search_name']] = search_name
                    process.append(tuple(row.values))
                    break
            else:
                non_process.append(tuple(row.values))

        process = [self.header] + process
        procces_df = pd.DataFrame.from_records(process, columns=self.header)
        search_names = procces_df[self.header[self.columns_dict['search_name']]].tolist()[1:]
        non_process = [self.header] + non_process
        non_process_df = pd.DataFrame.from_records(non_process, columns=self.header)

        self.__add_sheets_to_input(procces_df, non_process_df, input_df)
        return search_names

    def __add_only_status(self, values):
        wb = load_workbook(self.path)
        ws = wb.get_sheet_by_name(self.process_sheet)
        last_col = ws.max_column + 1
        for row, value in enumerate(values, start=1):
            ws.cell(row=row, column=last_col, value=value)
        wb.save(self.path)
        wb.close()

    def __rewrite_sheets_with_status(self, values):
        with pd.ExcelFile(self.path) as reader:
            input_df = pd.read_excel(reader, self.input_sheet, header=None, na_filter=False)
            procces_df = pd.read_excel(reader, self.process_sheet, header=None, na_filter=False)
            if self.has_non_processed:
                non_process_df = pd.read_excel(reader, self.no_process_sheet, header=None, na_filter=False)
        last_col = len(input_df.columns)
        procces_df.insert(last_col, column=str(last_col), value=values)
        with pd.ExcelWriter(self.path, mode='w') as writer:
            input_df.to_excel(writer, self.input_sheet, index=False, header=False)
            procces_df.to_excel(writer, self.process_sheet, index=False, header=False)
            if self.has_non_processed:
                non_process_df.to_excel(writer, self.no_process_sheet, index=False, header=False)

    def add_status_column(self, values):
        logger.info(self.config.get('info', 'add_status'))
        values = [self.config.get('settings', 'status_col')] + values
        try:
            try:
                self.__add_only_status(values)
            except Exception:  # if file is .xls
                self.__rewrite_sheets_with_status(values)
        except Exception as e:
            logger.error(self.config.get('error', 'add_status').format(path.basename(self.path)))
            raise e


class Output:

    def __init__(self, config):
        self.config = config
        self.path = path.join(config.get('path', 'temp'), config.get('settings', 'output_file').format(
            datetime.now().strftime(config.get('settings', 'file_date_format'))))

    def write(self, data_frame):
        logger.info(self.config.get('info', 'write_output'))
        try:
            with pd.ExcelWriter(self.path) as writer:
                data_frame.to_excel(writer, self.config.get('settings', 'output_sheet'), index=False, header=True)
        except Exception as e:
            logger.error(self.config.get('error', 'write_output').format(path.basename(self.path)))
            raise e
